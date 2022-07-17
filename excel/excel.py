'''
Code that helps interacting with an excel file
'''

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook

from excel.consts import EXCEL_FILENAME
from typing import List


def open_workbook(filename):
    wb = load_workbook(filename)
    return wb


def formatSubCategory(ws: Worksheet, row: int):
    """formatSubCategory takes a row and creates the basic subcategory format
    of that row. This means merging the cells from A to E (for now)"""
    left = 'A' + str(row)
    middle = 'E' + str(row)
    right = 'F' + str(row)

    rng = f"{left}:{middle}"

    # do not merge cells if they're already merged
    if rng in ws.merged_cells:
        return

    # This is necessary because merging throws an error if there already exists
    # some kind of merged cells in there
    cell_names = 'ABCDE'
    for i in range(len(cell_names) - 1):
        for j in range(i+1, len(cell_names)):
            cell_rng = f"{cell_names[i] + str(row)}:{cell_names[j] + str(row)}"
            if cell_rng in ws.merged_cells:
                ws.unmerge_cells(cell_rng)



    ws.merge_cells(f"{left}:{middle}")

def formatSubCategories(ws: Worksheet, rows: list[int]):
    for row in rows:
        formatSubCategory(ws, row)

# This is just some code to format my book nicer
all_rows = [i for i in range(29, 133)]

wb = open_workbook(EXCEL_FILENAME)
ws = wb.active

print('A127:E127' in ws.merged_cells)

print(wb.active)
formatSubCategories(wb.active, all_rows)
wb.save(EXCEL_FILENAME)
wb.close()

formatSubCategory(wb.active, 29)